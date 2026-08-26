"""운영 중인 근무시간표 엑셀에서 학생별 날짜 단위 가용시간을 추출.

정보서비스팀에서 실제 사용하는 워크북 구조를 파싱한다:
- `명단` 시트: 학생 목록, 재원(교비/국가), 근무 시작일, 시즌별 근무 여부
- 학생별 개인 시트: 주차 블록이 가로로 나열, 각 블록 = 라벨 열 + 요일 7열
  - 시트 전체가 r2=08:00 기준 30분/행 그리드이며, 슬롯은 셀 병합으로 표현됨
    (평일 템플릿, 토요일 2교대, 대청소 같은 임시 슬롯 모두 병합 범위가 다름)
  - 따라서 요일 열의 병합 범위(없으면 단일 셀)를 그대로 시간 구간으로 해석
  - 상태값: 근무 가능 / 근무 희망 / 수업 / 식사 / 기타 / 시험 / 대청소 / 빈칸

출력: 스케줄러 샘플 데이터 JSON (date_schedule 방식).

실행:
    cd backend
    .venv/bin/python -m app.scheduler.tools.import_xlsx "<엑셀 경로>" \
        --start 2026-03-02 --days 154 --out students_2026_1
"""

import argparse
import json
import re
from datetime import date, datetime, timedelta
from pathlib import Path

import openpyxl

STATUS_AVAILABLE = {"근무 가능"}
STATUS_PREFERRED = {"근무 희망"}
STATUS_CLASS = {"수업"}

DATE_HEADER_RE = re.compile(r"^(\d{1,2})\.(\d{1,2})\(")
WEEK_HEADER_RE = re.compile(r"주차")

GRID_BASE_ROW = 2  # r2 = 08:00
GRID_LAST_ROW = 29  # r29 = 21:30~22:00
GRID_BASE_MINUTE = 8 * 60
GRID_ROW_MINUTES = 30

# 개인 시트 공지의 팀 규칙: "오전 8시 근무는 격일 이상으로 배정"
DEFAULT_PREFERENCES = {
    "no_morning_after_close": True,
    "max_consecutive_morning_days": 1,
    "max_morning_days_per_week": 6,
    "wants_meal_break": False,  # 식사 시간은 학생이 시트에 직접 마킹(가용시간에서 제외됨)
}


def parse_roster(wb) -> list[dict]:
    ws = wb["명단"]
    header = [c.value for c in ws[1]]
    idx = {name: i for i, name in enumerate(header)}
    roster = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[idx["이름"]]:
            continue
        start_raw = str(row[idx["근무시작일"]]).replace(".", "-")[:10]
        roster.append(
            {
                "name": row[idx["이름"]],
                "funding": "gukga" if row[idx["그룹"]] == "국가" else "gyobi",
                "work_start": start_raw,
                "spring": bool(row[idx["봄"]]),
                "summer": bool(row[idx["여름"]]),
            }
        )
    return roster


def find_blocks(ws) -> list[tuple[int, list[tuple[int, date]]]]:
    """(라벨 열, [(요일 열, 날짜)]) 블록 목록."""
    blocks = []
    for cell in ws[1]:
        if not (isinstance(cell.value, str) and WEEK_HEADER_RE.search(cell.value)):
            continue
        label_col = cell.column
        days = []
        for offset in range(1, 8):
            raw = ws.cell(row=1, column=label_col + offset).value
            m = DATE_HEADER_RE.match(str(raw or ""))
            if not m:
                break
            month, dom = int(m.group(1)), int(m.group(2))
            days.append((label_col + offset, date(2026, month, dom)))
        if days:
            blocks.append((label_col, days))
    return blocks


def _row_to_minute(row: int) -> int:
    return GRID_BASE_MINUTE + (row - GRID_BASE_ROW) * GRID_ROW_MINUTES


def _minute_to_str(minute: int) -> str:
    return f"{minute // 60:02d}:{minute % 60:02d}"


def build_merge_map(ws) -> dict[tuple[int, int], tuple[int, int]]:
    """(행, 열) → 그 셀이 속한 병합 범위의 (시작 행, 끝 행)."""
    merge_map = {}
    for rng in ws.merged_cells.ranges:
        if rng.max_row < GRID_BASE_ROW or rng.min_row > GRID_LAST_ROW:
            continue
        for c in range(rng.min_col, rng.max_col + 1):
            for r in range(rng.min_row, rng.max_row + 1):
                merge_map[(r, c)] = (rng.min_row, rng.max_row)
    return merge_map


def parse_day_column(ws, merge_map, col: int) -> dict[str, list[list[str]]]:
    """요일 열 하나에서 {available, preferred, classes} 시간 구간 추출.

    시트는 r2=08:00 기준 30분/행 그리드이고 슬롯은 셀 병합으로 표현되므로,
    병합 범위(없으면 단일 셀)를 그대로 시간 구간으로 해석한다. 상태값은
    병합 앵커(좌상단) 셀에 있다.
    """
    out = {"available": [], "preferred": [], "classes": []}
    row = GRID_BASE_ROW
    while row <= GRID_LAST_ROW:
        top, bottom = merge_map.get((row, col), (row, row))
        bottom = min(bottom, GRID_LAST_ROW)
        status = ws.cell(row=top, column=col).value
        start = _minute_to_str(_row_to_minute(max(top, GRID_BASE_ROW)))
        end = _minute_to_str(_row_to_minute(bottom) + GRID_ROW_MINUTES)
        if isinstance(status, str):
            status = status.strip()
            if status in STATUS_PREFERRED:
                out["available"].append([start, end])
                out["preferred"].append([start, end])
            elif status in STATUS_AVAILABLE:
                out["available"].append([start, end])
            elif status in STATUS_CLASS:
                out["classes"].append([start, end])
        row = bottom + 1
    return out


def parse_student_sheet(ws, start: date, end: date) -> dict[str, dict]:
    """개인 시트 → {ISO날짜: {available, preferred, classes}}."""
    schedule = {}
    merge_map = build_merge_map(ws)
    for _, days in find_blocks(ws):
        for col, day in days:
            if not (start <= day <= end):
                continue
            parsed = parse_day_column(ws, merge_map, col)
            if any(parsed.values()):
                schedule[day.isoformat()] = parsed
    return schedule


def main() -> None:
    parser = argparse.ArgumentParser(description="근무시간표 엑셀 → 스케줄러 JSON")
    parser.add_argument("xlsx_path")
    parser.add_argument("--start", default="2026-03-02", help="기간 시작일 (월요일 권장)")
    parser.add_argument("--days", type=int, default=154, help="기간 일수")
    parser.add_argument("--out", default="students_2026_1", help="출력 샘플 이름")
    args = parser.parse_args()

    start = date.fromisoformat(args.start)
    end = start + timedelta(days=args.days - 1)
    wb = openpyxl.load_workbook(args.xlsx_path, data_only=True)

    # 여름학기 시작일: 봄만 근무하는 학생의 활동 종료 경계
    summer_start = date(2026, 6, 23)

    students = []
    for entry in parse_roster(wb):
        name = entry["name"]
        if name not in wb.sheetnames:
            print(f"경고: {name} 개인 시트 없음 — 건너뜀")
            continue
        if not (entry["spring"] or entry["summer"]):
            continue
        active_from = max(date.fromisoformat(entry["work_start"]), start)
        if not entry["spring"]:
            active_from = max(active_from, summer_start)
        active_until = end if entry["summer"] else summer_start - timedelta(days=1)

        schedule = parse_student_sheet(wb[name], start, end)
        n_days = len(schedule)
        students.append(
            {
                "student_id": name,
                "name": name,
                "funding_type": entry["funding"],
                "active_from": active_from.isoformat(),
                "active_until": active_until.isoformat(),
                "date_schedule": schedule,
                "preferences": dict(DEFAULT_PREFERENCES),
            }
        )
        print(f"{name}: {entry['funding']}, 활동 {active_from}~{active_until}, {n_days}일 파싱")

    out = {
        "_comment": f"'{Path(args.xlsx_path).name}'에서 자동 추출 ({datetime.now():%Y-%m-%d}). "
        "재추출: python -m app.scheduler.tools.import_xlsx",
        "horizon": {"start_date": start.isoformat(), "num_days": args.days},
        "students": students,
    }
    out_path = Path(__file__).parent.parent / "config" / "sample" / f"{args.out}.json"
    out_path.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n저장: {out_path} (학생 {len(students)}명)")


if __name__ == "__main__":
    main()
